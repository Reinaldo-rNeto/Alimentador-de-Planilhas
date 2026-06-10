# -*- coding: utf-8 -*-
import re
from datetime import date
from dataclasses import dataclass, field
from typing import Optional
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PRIORIDADES = ["Crítica", "Alta", "Média", "Baixa"]

CORES = {
    "Crítica": {"header": "FFB71C1C", "row": "FFFFEBEE", "font": "FFFFFFFF"},
    "Alta":    {"header": "FFE65100", "row": "FFFBE9E7", "font": "FFFFFFFF"},
    "Média":   {"header": "FF1565C0", "row": "FFE3F2FD", "font": "FFFFFFFF"},
    "Baixa":   {"header": "FF37474F", "row": "FFECEFF1", "font": "FFFFFFFF"},
}

# Hex sem alpha para uso no tkinter
CORES_UI = {
    "Crítica": {"bg": "#B71C1C", "row": "#FFEBEE", "fg": "#FFFFFF"},
    "Alta":    {"bg": "#E65100", "row": "#FBE9E7", "fg": "#FFFFFF"},
    "Média":   {"bg": "#1565C0", "row": "#E3F2FD", "fg": "#FFFFFF"},
    "Baixa":   {"bg": "#37474F", "row": "#ECEFF1", "fg": "#FFFFFF"},
}

# Valores exatos conforme validação da planilha original
STATUS_OPCOES = [
    "A Iniciar", "Discovery", "Lev. de Req", "Desenvovimento",
    "Suspenso", "Impeditivo", "Análise", "-",
    "Acompanhamento", "Homologação", "Concluído", "Em andamento",
]
IMPEDITIVO_OPCOES = ["Normal", "Atenção", "SIM", "-"]
TIPO_OPCOES = ["Projeto", "Operacional", "Dem. Pontual"]
NATUREZA_OPCOES = ["Externo", "Interno"]

# Mapeamento fixo de colunas (1-based)
COL = {
    "marcador":      1,
    "prio":          2,
    "num":           3,
    "projeto":       4,
    "natureza":      5,
    "demandante":    6,
    "dt_inicio":     7,
    "dt_estimada":   8,
    "dt_entrega":    9,
    "responsavel":   10,
    "acompanhamento":11,
    "tipo":          12,
    "status":        13,
    "impeditivo":    14,
    "observacao":    15,
    "notas":         16,
}

HEADER_ROW_OFFSET = 1   # linha do cabeçalho relativa ao início da seção


@dataclass
class Projeto:
    linha: int                     # linha real na planilha (0 = novo)
    secao: str                     # "backlog" ou "projetos"
    prio: str = "Alta"
    num: Optional[float] = None
    nome: str = ""
    natureza: str = "Externo"
    demandante: str = ""
    dt_inicio: str = "-"
    dt_estimada: str = "-"
    dt_entrega: str = "-"
    responsavel: str = ""
    acompanhamento: str = ""
    tipo: str = "Projeto"
    status: str = "A Iniciar"
    impeditivo: str = "Normal"
    observacao: str = ""
    notas: str = ""
    marcador: str = ""


@dataclass
class Secao:
    titulo: str
    header_row: int
    data_start: int
    data_end: int
    grupos: dict = field(default_factory=dict)       # prio -> [linhas de projetos]
    separadores: dict = field(default_factory=dict)  # prio -> linha do separador


@dataclass
class GenericRow:
    """Linha de planilha genérica (modo não-estruturado)."""
    linha: int
    dados: dict   # nome_coluna -> valor


class SheetManager:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = None
        self.ws = None
        self.sheet_name = None
        self.secoes: list[Secao] = []
        self.projetos: list[Projeto] = []
        # Modo genérico
        self.mode: str = "structured"        # "structured" | "generic"
        self.generic_headers: list[tuple] = []  # [(col_idx, nome), ...]
        self.generic_rows: list[GenericRow] = []
        self.generic_header_row: int = 1

    # ------------------------------------------------------------------ #
    #  Carregamento                                                        #
    # ------------------------------------------------------------------ #

    def load(self, sheet_name: str = None):
        self.wb = openpyxl.load_workbook(self.filepath)
        names = self.wb.sheetnames
        if sheet_name and sheet_name in names:
            self.sheet_name = sheet_name
        else:
            self.sheet_name = names[0]
        self.ws = self.wb[self.sheet_name]
        self._parse()

    def get_sheet_names(self) -> list[str]:
        if not self.wb:
            wb = openpyxl.load_workbook(self.filepath, read_only=True)
            names = wb.sheetnames
            wb.close()
            return names
        return self.wb.sheetnames

    # ------------------------------------------------------------------ #
    #  Parse da planilha                                                   #
    # ------------------------------------------------------------------ #

    def _cell_val(self, row: int, col: int):
        v = self.ws.cell(row=row, column=col).value
        if v is None:
            return ""
        if hasattr(v, "strftime"):
            return v.strftime("%d/%m/%Y")
        return str(v).strip()

    def _parse(self):
        self.secoes = []
        self.projetos = []
        self.generic_rows = []
        self.generic_headers = []

        header_rows = self._find_header_rows()
        if header_rows:
            self.mode = "structured"
            for i, h in enumerate(header_rows):
                next_h = header_rows[i + 1] if i + 1 < len(header_rows) else None
                secao = self._parse_secao(h, next_h)
                self.secoes.append(secao)
        else:
            self.mode = "generic"
            self._parse_generic()

    def _find_header_rows(self) -> list[int]:
        """Encontra linhas com os cabeçalhos 'Prio', '#', 'Projeto'."""
        found = []
        for r in range(1, self.ws.max_row + 1):
            v = self.ws.cell(row=r, column=COL["prio"]).value
            v2 = self.ws.cell(row=r, column=COL["num"]).value
            if str(v).strip() == "Prio" and str(v2).strip() == "#":
                found.append(r)
        return found

    # ------------------------------------------------------------------ #
    #  Parser genérico (qualquer planilha)                                 #
    # ------------------------------------------------------------------ #

    def _parse_generic(self):
        """Detecta header e lê todas as linhas como tabela plana."""
        max_col = self.ws.max_column
        header_row = None

        # Procura a primeira linha com ≥ 2 células não-vazias nos primeiros 20 rows
        for r in range(1, min(21, self.ws.max_row + 1)):
            vals = [self.ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            non_empty = [(i + 1, v) for i, v in enumerate(vals)
                         if v is not None and str(v).strip()]
            if len(non_empty) >= 2:
                header_row = r
                # Usa os valores não-nulos como cabeçalhos; duplicados recebem sufixo
                seen: dict[str, int] = {}
                hdrs = []
                for col_idx, val in non_empty:
                    name = str(val).strip()
                    if name in seen:
                        seen[name] += 1
                        name = f"{name}_{seen[name]}"
                    else:
                        seen[name] = 0
                    hdrs.append((col_idx, name))
                self.generic_headers = hdrs
                self.generic_header_row = r
                break

        if not header_row:
            return

        col_map = {col_idx: name for col_idx, name in self.generic_headers}

        for r in range(header_row + 1, self.ws.max_row + 1):
            row_vals = {}
            has_data = False
            for col_idx, name in self.generic_headers:
                v = self.ws.cell(row=r, column=col_idx).value
                if v is None:
                    row_vals[name] = ""
                else:
                    has_data = True
                    if hasattr(v, "strftime"):
                        row_vals[name] = v.strftime("%d/%m/%Y")
                    else:
                        row_vals[name] = str(v).strip()
            if has_data:
                self.generic_rows.append(GenericRow(linha=r, dados=row_vals))

    # ------------------------------------------------------------------ #
    #  Salvar linha genérica                                               #
    # ------------------------------------------------------------------ #

    def salvar_generic_row(self, grow: GenericRow):
        for col_idx, name in self.generic_headers:
            self.ws.cell(row=grow.linha, column=col_idx).value = grow.dados.get(name, "")

    def criar_generic_row(self, dados: dict) -> GenericRow:
        """Adiciona nova linha no final dos dados genéricos."""
        last = self.generic_rows[-1].linha if self.generic_rows else self.generic_header_row
        new_row = last + 1
        self.ws.insert_rows(new_row)
        grow = GenericRow(linha=new_row, dados=dados)
        for col_idx, name in self.generic_headers:
            self.ws.cell(row=new_row, column=col_idx).value = dados.get(name, "")
        self.generic_rows.append(grow)
        return grow

    def _parse_secao(self, header_row: int, next_header_row: Optional[int] = None) -> Secao:
        # Título: linha não-vazia mais próxima acima do cabeçalho
        titulo = ""
        for delta in range(1, 8):
            row = header_row - delta
            if row < 1:
                break
            for c in range(1, self.ws.max_column + 1):
                v = self.ws.cell(row=row, column=c).value
                if v and str(v).strip():
                    titulo = str(v).strip()
                    break
            if titulo:
                break

        data_start = header_row + 1
        # Limite: linha anterior ao próximo cabeçalho ou fim da planilha
        hard_limit = (next_header_row - 1) if next_header_row else self.ws.max_row

        # Encontrar última linha não-vazia dentro do limite
        max_data = data_start
        for r in range(data_start, hard_limit + 1):
            vals = [self.ws.cell(row=r, column=c).value for c in range(2, 16)]
            if any(v is not None and str(v).strip() not in ("", "None") for v in vals):
                max_data = r

        secao = Secao(
            titulo=titulo,
            header_row=header_row,
            data_start=data_start,
            data_end=max_data,
            grupos={p: [] for p in PRIORIDADES},
            separadores={},
        )

        prio_atual = "Alta"  # default razoável
        for r in range(data_start, max_data + 1):
            val_prio = str(self.ws.cell(row=r, column=COL["prio"]).value or "").strip()
            val_num  = self.ws.cell(row=r, column=COL["num"]).value
            nome     = self._cell_val(r, COL["projeto"])

            # Separador: ALL CAPS (ex: "ALTA", " CRÍTICO") sem número associado
            upper = val_prio.upper()
            stripped = val_prio.strip()
            is_all_caps = stripped == stripped.upper() and stripped != ""
            is_separator = is_all_caps and (
                upper.strip() in [p.upper() for p in PRIORIDADES]
                or "CRÍTICO" in upper
                or "CRITICO" in upper
            ) and (val_num is None or not isinstance(val_num, (int, float)))

            if is_separator:
                for p in PRIORIDADES:
                    if p.upper() in upper or ("CRÍTICO" in upper and p == "Crítica"):
                        prio_atual = p
                        secao.separadores[p] = r   # rastreia a linha do separador
                        break
                continue

            # Linha com prioridade explícita na célula
            if val_prio in PRIORIDADES:
                prio_atual = val_prio

            # Precisamos de pelo menos nome ou número para considerar projeto válido
            if not nome and not isinstance(val_num, (int, float)):
                continue

            proj = Projeto(
                linha=r,
                secao=secao.titulo,
                prio=prio_atual,
                num=float(val_num) if isinstance(val_num, (int, float)) else None,
                nome=nome,
                natureza=self._cell_val(r, COL["natureza"]),
                demandante=self._cell_val(r, COL["demandante"]),
                dt_inicio=self._cell_val(r, COL["dt_inicio"]),
                dt_estimada=self._cell_val(r, COL["dt_estimada"]),
                dt_entrega=self._cell_val(r, COL["dt_entrega"]),
                responsavel=self._cell_val(r, COL["responsavel"]),
                acompanhamento=self._cell_val(r, COL["acompanhamento"]),
                tipo=self._cell_val(r, COL["tipo"]),
                status=self._cell_val(r, COL["status"]),
                impeditivo=self._cell_val(r, COL["impeditivo"]),
                observacao=self._cell_val(r, COL["observacao"]),
                notas=self._cell_val(r, COL["notas"]),
                marcador=self._cell_val(r, COL["marcador"]),
            )
            self.projetos.append(proj)
            secao.grupos.setdefault(prio_atual, []).append(r)

        return secao

    # ------------------------------------------------------------------ #
    #  Adicionar observação                                                #
    # ------------------------------------------------------------------ #

    def adicionar_observacao(self, projeto: Projeto, texto: str):
        hoje = date.today().strftime("%d/%m")
        nova_linha = f"- {hoje} {texto.strip()}"
        obs_atual = self.ws.cell(row=projeto.linha, column=COL["observacao"]).value or ""
        if obs_atual:
            nova_obs = nova_linha + "\n" + str(obs_atual)
        else:
            nova_obs = nova_linha
        self.ws.cell(row=projeto.linha, column=COL["observacao"]).value = nova_obs
        projeto.observacao = nova_obs

    # ------------------------------------------------------------------ #
    #  Atualizar campos de um projeto existente                           #
    # ------------------------------------------------------------------ #

    def salvar_projeto(self, projeto: Projeto):
        r = projeto.linha
        self.ws.cell(row=r, column=COL["prio"]).value = projeto.prio
        self.ws.cell(row=r, column=COL["projeto"]).value = projeto.nome
        self.ws.cell(row=r, column=COL["natureza"]).value = projeto.natureza
        self.ws.cell(row=r, column=COL["demandante"]).value = projeto.demandante
        self.ws.cell(row=r, column=COL["dt_inicio"]).value = projeto.dt_inicio
        self.ws.cell(row=r, column=COL["dt_estimada"]).value = projeto.dt_estimada
        self.ws.cell(row=r, column=COL["dt_entrega"]).value = projeto.dt_entrega
        self.ws.cell(row=r, column=COL["responsavel"]).value = projeto.responsavel
        self.ws.cell(row=r, column=COL["acompanhamento"]).value = projeto.acompanhamento
        self.ws.cell(row=r, column=COL["tipo"]).value = projeto.tipo
        self.ws.cell(row=r, column=COL["status"]).value = projeto.status
        self.ws.cell(row=r, column=COL["impeditivo"]).value = projeto.impeditivo
        self.ws.cell(row=r, column=COL["observacao"]).value = projeto.observacao
        self.ws.cell(row=r, column=COL["notas"]).value = projeto.notas
        self._aplicar_cores_linha(r, projeto.prio)

    # ------------------------------------------------------------------ #
    #  Criar novo projeto                                                  #
    # ------------------------------------------------------------------ #

    def criar_projeto(self, projeto: Projeto) -> Projeto:
        secao = self._get_secao_by_titulo(projeto.secao)
        if not secao:
            secao = self.secoes[0]

        # Encontrar última linha do grupo de prioridade ou último separador
        insert_row = self._encontrar_linha_insercao(secao, projeto.prio)
        self.ws.insert_rows(insert_row)

        # Atribuir número sequencial
        nums = [p.num for p in self.projetos if p.num is not None]
        projeto.num = (max(nums) + 1) if nums else 1

        projeto.linha = insert_row
        self.ws.cell(row=insert_row, column=COL["prio"]).value = projeto.prio
        self.ws.cell(row=insert_row, column=COL["num"]).value = projeto.num
        self.ws.cell(row=insert_row, column=COL["projeto"]).value = projeto.nome
        self.ws.cell(row=insert_row, column=COL["natureza"]).value = projeto.natureza
        self.ws.cell(row=insert_row, column=COL["demandante"]).value = projeto.demandante
        self.ws.cell(row=insert_row, column=COL["dt_inicio"]).value = projeto.dt_inicio
        self.ws.cell(row=insert_row, column=COL["dt_estimada"]).value = projeto.dt_estimada
        self.ws.cell(row=insert_row, column=COL["dt_entrega"]).value = projeto.dt_entrega
        self.ws.cell(row=insert_row, column=COL["responsavel"]).value = projeto.responsavel
        self.ws.cell(row=insert_row, column=COL["acompanhamento"]).value = projeto.acompanhamento
        self.ws.cell(row=insert_row, column=COL["tipo"]).value = projeto.tipo
        self.ws.cell(row=insert_row, column=COL["status"]).value = projeto.status
        self.ws.cell(row=insert_row, column=COL["impeditivo"]).value = projeto.impeditivo
        self.ws.cell(row=insert_row, column=COL["observacao"]).value = projeto.observacao

        self._aplicar_cores_linha(insert_row, projeto.prio)
        self._aplicar_validacao(insert_row)
        self.projetos.append(projeto)
        return projeto

    def _encontrar_linha_insercao(self, secao: Secao, prio: str) -> int:
        linhas_do_grupo = secao.grupos.get(prio, [])
        if linhas_do_grupo:
            return max(linhas_do_grupo) + 1
        # Grupo vazio: inserir logo após o separador da prioridade
        sep_row = secao.separadores.get(prio)
        if sep_row:
            return sep_row + 1
        # Sem separador: inserir no fim da seção
        return secao.data_end + 1

    def _get_secao_by_titulo(self, titulo: str) -> Optional[Secao]:
        for s in self.secoes:
            if s.titulo == titulo:
                return s
        return None

    # ------------------------------------------------------------------ #
    #  Validação de dropdown                                               #
    # ------------------------------------------------------------------ #

    # Fórmulas exatas da planilha original (com espaços/typos preservados)
    _VALIDACOES = {
        COL["natureza"]:   '"Interno,Externo"',
        COL["tipo"]:       '"Projeto,Operacional,Dem. Pontual"',
        COL["status"]:     '"A Iniciar,Discovery ,Lev. de Req,Desenvovimento,Suspenso,Impeditivo,Análise,-,Acompanhamento,Homologação,Concluído ,Em andamento"',
        COL["impeditivo"]: '"SIM,Atenção,Normal,-"',
    }

    def _aplicar_validacao(self, row: int):
        from openpyxl.worksheet.datavalidation import DataValidation
        for col, formula in self._VALIDACOES.items():
            cell_ref = f"{get_column_letter(col)}{row}"
            # Reutilizar validação existente da mesma fórmula
            found = False
            for dv in self.ws.data_validations.dataValidation:
                if dv.type == "list" and dv.formula1 == formula:
                    dv.add(cell_ref)
                    found = True
                    break
            if not found:
                dv = DataValidation(type="list", formula1=formula, showDropDown=False)
                dv.sqref = cell_ref
                self.ws.add_data_validation(dv)

    # ------------------------------------------------------------------ #
    #  Criar nova seção (grupo separador + linhas de prioridade)           #
    # ------------------------------------------------------------------ #

    def criar_secao(self, titulo: str, insert_after_row: int = None) -> Secao:
        """Insere um bloco de nova seção na planilha com cabeçalho e separadores."""
        from openpyxl.styles import PatternFill, Font, Alignment

        # Onde inserir: após a última seção existente ou posição indicada
        if insert_after_row is None:
            last = self.secoes[-1] if self.secoes else None
            insert_after_row = (last.data_end + 2) if last else 10

        base = insert_after_row + 1   # linha do título/banner

        # Inserir linhas: título + cabeçalho + 4 separadores (um por prioridade) + linha vazia
        total_linhas = 2 + len(PRIORIDADES)
        self.ws.insert_rows(base, amount=total_linhas)

        # --- Linha do título da seção ---
        titulo_fill  = PatternFill("solid", fgColor="FF1A237E")
        titulo_font  = Font(bold=True, color="FFFFFFFF", size=13)
        titulo_align = Alignment(vertical="center", horizontal="left")
        cell_titulo  = self.ws.cell(row=base, column=COL["prio"])
        cell_titulo.value     = titulo
        cell_titulo.fill      = titulo_fill
        cell_titulo.font      = titulo_font
        cell_titulo.alignment = titulo_align
        # Mesclar visualmente (merge simples nos dados da coluna D)
        self.ws.cell(row=base, column=COL["projeto"]).value = ""

        # --- Linha do cabeçalho ---
        hdr_row  = base + 1
        hdr_fill = PatternFill("solid", fgColor="FF0D47A1")
        hdr_font = Font(bold=True, color="FFFFFFFF")
        colunas_hdr = {
            COL["prio"]:          "Prio",
            COL["num"]:           "#",
            COL["projeto"]:       "Projeto",
            COL["natureza"]:      "Natureza",
            COL["demandante"]:    "Demandante",
            COL["dt_inicio"]:     "DT Início",
            COL["dt_estimada"]:   "DT Estimada",
            COL["dt_entrega"]:    "DT Entrega",
            COL["responsavel"]:   "Responsável",
            COL["acompanhamento"]:"Acompanhamento",
            COL["tipo"]:          "Tipo",
            COL["status"]:        "Status de andamento",
            COL["impeditivo"]:    "Impeditivo",
            COL["observacao"]:    "Observação",
        }
        for col, label in colunas_hdr.items():
            c = self.ws.cell(row=hdr_row, column=col)
            c.value = label
            c.fill  = hdr_fill
            c.font  = hdr_font

        # --- Separadores por prioridade ---
        sep_nomes = {
            "Crítica": "CRÍTICO",
            "Alta":    "ALTA",
            "Média":   "MÉDIA",
            "Baixa":   "BAIXA",
        }
        secao = Secao(
            titulo=titulo,
            header_row=hdr_row,
            data_start=hdr_row + 1,
            data_end=hdr_row + len(PRIORIDADES),
            grupos={p: [] for p in PRIORIDADES},
            separadores={},
        )

        for i, prio in enumerate(PRIORIDADES):
            sep_row = hdr_row + 1 + i
            sep_fill = PatternFill("solid", fgColor=CORES[prio]["header"])
            sep_font = Font(bold=True, color=CORES[prio]["font"])
            c = self.ws.cell(row=sep_row, column=COL["prio"])
            c.value = sep_nomes[prio]
            c.fill  = sep_fill
            c.font  = sep_font
            secao.separadores[prio] = sep_row

        self.secoes.append(secao)
        self.wb.save(self.filepath)
        return secao

    # ------------------------------------------------------------------ #
    #  Cores                                                               #
    # ------------------------------------------------------------------ #

    def _aplicar_cores_linha(self, row: int, prio: str):
        if prio not in CORES:
            return
        header_fill = PatternFill("solid", fgColor=CORES[prio]["header"])
        row_fill = PatternFill("solid", fgColor=CORES[prio]["row"])
        header_font = Font(bold=True, color=CORES[prio]["font"])
        row_font = Font(color="FF000000")

        # Col prio: cor cheia com texto branco bold
        cell_prio = self.ws.cell(row=row, column=COL["prio"])
        cell_prio.fill = header_fill
        cell_prio.font = header_font

        # Demais colunas: fundo claro
        for c in range(COL["num"], COL["notas"] + 1):
            cell = self.ws.cell(row=row, column=c)
            cell.fill = row_fill
            cell.font = row_font

    # ------------------------------------------------------------------ #
    #  Salvar arquivo                                                      #
    # ------------------------------------------------------------------ #

    def save(self, filepath: str = None):
        target = filepath or self.filepath
        self.wb.save(target)

    def reload(self, sheet_name: str = None):
        self.load(sheet_name or self.sheet_name)
